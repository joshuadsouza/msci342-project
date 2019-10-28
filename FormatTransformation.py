from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM


def transform_file(filepath):
    MainBook = load_workbook(filepath)

    for sheet in MainBook:

        #Bulk delete
        #Remove columns A-M
        sheet.delete_cols(1,12)

        #Remove first row
        sheet.delete_rows(1,1)

    #Add day dividers
        testValue = ""
        i = 1
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row*2):
            for cell in row:
                if cell.value:
                    if cell.value != testValue:
                        sheet.insert_rows(i)
                        hold = sheet.cell(row = i,column = 2, value = cell.value)
                        hold.alignment = Alignment(horizontal='left')
            testValue = cell.value
            i += 1

        #Borders
        #Delete column A
        sheet.delete_cols(1,1)
        #Add Border to all active cells
        #Create Style
        med_border = Border(
        left=Side(border_style=BORDER_MEDIUM, color='00000000'),
        right=Side(border_style=BORDER_MEDIUM, color='00000000'),
        top=Side(border_style=BORDER_MEDIUM, color='00000000'),
        bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
        )
        #Add borders
        k=1
        for row in sheet.iter_rows(min_row=1, max_col=2, max_row=sheet.max_row):
            for cell in row:
                if cell.value:
                    sheet.cell(row=k,column=1).border = med_border
                    sheet.cell(row=k,column=2).border = med_border
            k+=1

        #Add Title
        #Create Font
        Font1 = Font(name='Calibri',
                    size=36,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

        sheet.insert_rows(1)
        sheet.row_dimensions[1].height = 45
        hold = sheet.cell(row = 1,column = 1, value = "First -Year Orientation")
        sheet.cell(row=1,column=1).font= Font1

        #Add Housing Group
        sheet.insert_rows(2)
        if sheet.title == "Architecture":
            House2 = "Architecture"
        else:
            House1 = sheet.title
            House1.split(' ', 1)
            House2 = House1.split(' ', 1)[1]

        hold = sheet.cell(row = 2,column = 1, value = House2)
        sheet.insert_rows(2)

        #Add Edge columns and format dimensions
        sheet.insert_cols(1)
        sheet.column_dimensions['A'].width = 3
        sheet.column_dimensions['B'].width = 135
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 3

    MainBook.save(filepath)
    return filepath